from openpyxl import Workbook
from random import randrange
from itertools import product
import os


def fill_cell():
    # some calculations
    return randrange(332)


def fill_sheet(workbook_name, max_row, max_col):
    for row, column in product(range(1, max_row + 1), range(1, max_col + 1)):
        workbook_name.active.cell(row=row, column=column).value = fill_cell()


def create_sheets(workbook_name, sheet_name, number_sheets):
    for sheet in range(1, number_sheets + 1):
        workbook_name.create_sheet(str(sheet_name) + "sheet_" + str(sheet))


def main():
    number_input = input("Enter number of files: ")
    print(f"Creating: {number_input} files...")
    file = 0;

    for file in range(int(number_input)):
        workbook = Workbook()
        active_worksheet = workbook.active
        fill_sheet(workbook, 4, 5)
        create_sheets(workbook, f"file_{file + 1}_", 2)
        workbook.save(f"./input/test_workbook_pr{file + 1}.xlsx")
        print(os.listdir('.'))


if __name__ == "__main__":
    main()
