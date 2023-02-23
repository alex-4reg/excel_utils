from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_key_word(filename, max_row, max_col, key, use_load_workbook=True):
    workbook = load_workbook(filename) if use_load_workbook else filename

    matches_coordinates = []
    for row in range(1, max_row + 1):
        for column in range(1, max_col + 1):
            cell_coordinate = get_column_letter(column) + str(row)
            if key in str(workbook.active[cell_coordinate].value):
                matches_coordinates.append(cell_coordinate)
    return matches_coordinates


def preprocessor(file):
    file_output = file.replace("input/", "output/processed_")
    source_workbook = load_workbook(file)

    for i in range(len(source_workbook.sheetnames)):
        source_workbook.active = source_workbook[source_workbook.sheetnames[i]]

        invoice_coordinate = find_key_word(source_workbook, 100, 3, "invoice", use_load_workbook=False)[0]
        invoice_cell_to_split = source_workbook.active[invoice_coordinate].value.split(". ")
        col_invoice_title, cell_value_invoice = invoice_cell_to_split[0], invoice_cell_to_split[1]
        date_cell = source_workbook.active.cell(source_workbook.active[invoice_coordinate].row,
                                                source_workbook.active[invoice_coordinate].column + 2).value
        date_cell = str(date_cell).rstrip(" 00:00:00")
        start_table_coordinate = find_key_word(source_workbook, 100, 3, "name", use_load_workbook=False)[0]
        source_workbook.active.cell(int(start_table_coordinate[1]),
                                    source_workbook.active.max_column + 1).value = col_invoice_title
        source_workbook.active.cell(int(start_table_coordinate[1]) + 1,
                                    source_workbook.active.max_column).value = cell_value_invoice
        source_workbook.active.cell(int(start_table_coordinate[1]),
                                    source_workbook.active.max_column + 1).value = "date"
        source_workbook.active.cell(int(start_table_coordinate[1]) + 1,
                                    source_workbook.active.max_column).value = date_cell

    source_workbook.save(filename=file_output)
    return file_output


if __name__ == "__main__":
    print(preprocessor("./input/materials.xlsx"))
