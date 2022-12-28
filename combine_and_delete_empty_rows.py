import openpyxl


def delete(current_sheet):
    for row in current_sheet.iter_rows():
        if all(not cell.value for cell in row):
            current_sheet.delete_rows(row[0].row, 1)
            delete(current_sheet)
            return


def combine(current_sheet, parent_column, dependent_column, additional_column):
    for i in range(1, current_sheet.max_row + 1):

        if current_sheet[parent_column + str(i)].value is not None:
            parent_cell = i
            if current_sheet[additional_column + str(i)].value is not None:
                parent_additional_cell = i

        elif current_sheet[parent_column + str(i)].value is None and current_sheet[dependent_column +
                                                                                   str(i)].value is not None:
            render_cell(current_sheet, parent_cell, i, dependent_column)
            # print(current_sheet[parent_column + str(i)].value)

        if current_sheet[additional_column + str(i)].value is not None and \
                current_sheet[additional_column + str(parent_additional_cell)].value is not None and \
                current_sheet[parent_column + str(i)].value is None:
            render_cell(current_sheet, parent_additional_cell, i, additional_column)


def render_cell(current_sheet, parent_index, current_index, *cols):
    for col in cols:
        current_sheet[col + str(parent_index)].value = current_sheet[col + str(parent_index)].value \
                                                       + " " + current_sheet[col + str(current_index)].value
        current_sheet[col + str(current_index)].value = None


if __name__ == '__main__':

    file = "./input/combine and delete empty rows.xlsx"
    work_book = openpyxl.load_workbook(file)
    sheet = work_book.active
    print("Rows before removing:", sheet.max_row)
    file = './output/combine and delete empty rows.xlsx'
    work_book.save(file)
    combine(sheet, "C", "A", "D")

    for row in sheet:
        delete(sheet)

    print("Rows after removing:", sheet.max_row)
    work_book.save(file)
