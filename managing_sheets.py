import os, fnmatch
from openpyxl import load_workbook

input_directory_prefix = './input'
output_directory_prefix = './output'
names_set = set()

listOfFiles = os.listdir(input_directory_prefix)
pattern = "*.xlsx"

listOfXlsxFiles = [entry for entry in listOfFiles if fnmatch.fnmatch(entry, pattern)]
destination_workbook = load_workbook(filename=input_directory_prefix + "/" + listOfXlsxFiles[0])
merged_file_name = listOfXlsxFiles[0].strip(".xlsx")
names_set.update(destination_workbook.sheetnames)
print(names_set)

for file in range(1, len(listOfXlsxFiles)):
    source_workbook = load_workbook(filename=input_directory_prefix + "/" + listOfXlsxFiles[file])
    merged_file_name = merged_file_name + "_" + listOfXlsxFiles[file].strip(".xlsx")

    for i in range(len(source_workbook.sheetnames)):
        current_sheet = source_workbook[source_workbook.sheetnames[i]]
        source_workbook.active = current_sheet

        if not names_set.__contains__(source_workbook.sheetnames[i]):
            names_set.add(source_workbook.sheetnames[i])

        else:
            source_workbook.active.title = str(file + 1) + "." + (source_workbook.sheetnames[i])
            current_sheet = source_workbook[source_workbook.sheetnames[i]]
            names_set.add(source_workbook.sheetnames[i])

        source_workbook.active._parent = destination_workbook
        print(source_workbook.active)
        destination_workbook._add_sheet(source_workbook.active)

    if file == len(listOfXlsxFiles) - 1:
        destination_workbook.save(filename=output_directory_prefix + "/merged_" + merged_file_name + ".xlsx")
